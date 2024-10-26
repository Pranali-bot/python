import openpyxl
from openpyxl import Workbook
import os


filename = "data.xlsx"


def initialize_excel_file():
    if not os.path.exists(filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "data"
        sheet.append(["Name", "Email", "Phone Number"])
        workbook.save(filename)


def add_user():
    name = input("Enter the Name: ")
    email = input("Enter the  Email id: ")
    phone = input("Enter phone Number: ")
    
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.append([name, email, phone])
    workbook.save(filename)
    print("User added successfully.\n")


def display_users():
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    print("\nStored Users:")
    print("-" * 30)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Name: {row[0]}, Email: {row[1]}, Phone: {row[2]}")
    print("-" * 30 + "\n")

# Main function
def main():
    initialize_excel_file()
    
    while True:
        print(" make a Choice to proceed:")
        print("1. Add User")
        print("2. Display Users")
        print("3. Exit")
        
        choice = input("press key to continue : ")
        
        if choice == "1":
            add_user()
        elif choice == "2":
            display_users()
        elif choice == "3":
            print("Exiting the program.")
            break
        else:
            print("Invalid choice. Please try again.\n")

# Run  main function
if __name__ == "__main__":
    main()